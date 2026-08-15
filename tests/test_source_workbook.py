from __future__ import annotations

import hashlib
import io
import json
import zipfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from flagella_repro.source_workbook import (
    DEPOSIT_NAME,
    DEPOSITED_TABLE_PATHS,
    FIGURE_LABELS,
    ORIGIN_DERIVED,
    ORIGIN_FILE,
    RESERVED_SHEETS,
    _sheet_name,
    _table_name,
    build_source_data_files,
    build_source_workbook,
    column_unit,
)


def _write_panel(
    root: Path,
    panel_id: str,
    tables: list[tuple[str, pd.DataFrame]],
    parameters: dict | None = None,
    seeds: dict | None = None,
) -> None:
    """Write one panel's tables and the matching provenance record.

    Each entry of ``tables`` is a repository-relative path and its content.
    """
    figure = panel_id.split("_")[0]
    outputs = []
    for relative, frame in tables:
        source = root / relative
        source.parent.mkdir(parents=True, exist_ok=True)
        if relative.endswith(".gz"):
            frame.to_csv(source, index=False, compression={"method": "gzip", "mtime": 0})
        else:
            frame.to_csv(source, index=False)
        outputs.append(
            {
                "relative_path": relative,
                "sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
                "bytes": source.stat().st_size,
                "rows": len(frame),
            }
        )
    provenance = {
        "schema_version": "1.0.0",
        "panel_id": panel_id,
        "status": "partial_reproduction",
        "generated_at_utc": "2026-08-11T00:00:00Z",
        "command": ["python", "plot.py"],
        "inputs": [],
        "outputs": outputs,
        "software": {"python": "3.12.11"},
        "parameters": parameters or {},
        "random_seeds": seeds or {},
        "limitations": ["Test fixture begins from processed data."],
    }
    record = root / "metadata" / "provenance" / figure.lower() / f"{panel_id}.json"
    record.parent.mkdir(parents=True, exist_ok=True)
    record.write_text(json.dumps(provenance), encoding="utf-8")


@pytest.fixture
def source_project(tmp_path: Path) -> Path:
    """A repository with three panels, one of which carries a legacy duplicate.

    Panel F2_A mirrors the real defect: a pre-renumbering copy under the retired
    panel name F2_X, and the current copy with three more columns.
    """
    _write_panel(
        tmp_path,
        "F1_C",
        [
            (
                "build/source_data/Figure_1/C/F1_C_distribution.csv",
                pd.DataFrame({"condition": ["WT", "mutant"], "speed_um_s": [1.0, 2.0]}),
            ),
            (
                "build/statistics/Figure_1/C/F1_C_statistics.csv",
                pd.DataFrame({"comparison": ["WT vs mutant"], "p_value": [0.01]}),
            ),
        ],
    )
    _write_panel(
        tmp_path,
        "F2_A",
        [
            (
                "data/source_data/figure_02/F2_X.csv",
                pd.DataFrame({"cell_id": [1, 2], "x_um": [3.5, 4.5]}),
            ),
            (
                "build/source_data/Figure_2/A/F2_A_source_data.csv",
                pd.DataFrame(
                    {
                        "cell_id": [1, 2],
                        "x_um": [3.5, 4.5],
                        "growth_rate_per_h": [0.4, 0.2],
                    }
                ),
            ),
        ],
    )
    _write_panel(
        tmp_path,
        "S3_B",
        [
            (
                "data/source_data/supplementary_03/S3_B_paired_points.csv",
                pd.DataFrame({"condition": ["WT", "mutant"], "mass_fraction": [0.4, 0.2]}),
            )
        ],
    )
    _write_deposit_panel(tmp_path, "S4_A")
    return tmp_path


def _trajectory(cells: int = 2, steps: int = 3) -> pd.DataFrame:
    """A miniature trajectory table with the simulator's real columns."""
    rows = []
    for cell in range(cells):
        for step in range(steps):
            rows.append(
                {
                    "phenotype": "WT",
                    "medium": "agarose",
                    "seed": 17,
                    "cell_id": cell,
                    "step": step,
                    "time_s": step * 0.5,
                    "x_um": float(step * (cell + 1)),
                    "y_um": 0.0,
                    "state": 0 if cell == 0 else 3,
                    "is_motile": cell == 0,
                }
            )
    return pd.DataFrame(rows)


def _write_deposit_panel(root: Path, panel_id: str) -> None:
    """Write a panel whose trajectory table is routed to the deposit."""
    trajectory_path = next(
        path for path in sorted(DEPOSITED_TABLE_PATHS) if f"/{panel_id}_" in path
    )
    _write_panel(
        root,
        panel_id,
        [
            (trajectory_path, _trajectory()),
            (
                "data/source_data/supplementary_04/S4_A_obstacles.csv",
                pd.DataFrame(
                    {
                        "phenotype": ["WT"],
                        "medium": ["agarose"],
                        "seed": [17],
                        "obstacle_index": [0],
                        "x_um": [10.0],
                        "y_um": [10.0],
                        "radius_um": [2.0],
                    }
                ),
            ),
        ],
        parameters={
            "phenotype": "WT",
            "medium": "agarose",
            "box_width_um": 148.0,
            "box_height_um": 96.0,
            "dt_s": 0.5,
            "track_duration_s": 1.0,
            "n_cells": 2,
            "obstacle_config": {
                "count": 1,
                "radius_range_um": [2.0, 2.0],
                "clearance_um": 1.7,
            },
            "motility_parameters": {
                "motile_fraction": 0.5,
                "run_speed_um_s": 19.9,
                "reorientation_rate_s": 6.2,
                "rotational_diffusion_rad2_s": 6.1,
                "turn_angle_sd_rad": 1.2,
                "passive_diffusion_um2_s": 0.35,
                "stall_probability": 0.0,
                "stall_mean_duration_s": 0.05,
            },
        },
        seeds={
            "panel_seed": 17,
            "starting_position_seed": 18,
            "obstacle_seed": 317,
        },
    )


def test_partial_source_workbook_is_deterministic_and_audited(tmp_path: Path) -> None:
    source = tmp_path / "data" / "source_data" / "figure_01" / "F1_C.csv"
    source.parent.mkdir(parents=True)
    pd.DataFrame({"condition": ["WT", "mutant"], "value": [1.0, 2.0]}).to_csv(
        source, index=False
    )
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    provenance = {
        "schema_version": "1.0.0",
        "panel_id": "F1_C",
        "status": "partial_reproduction",
        "generated_at_utc": "2026-08-11T00:00:00Z",
        "command": ["python", "plot.py"],
        "inputs": [],
        "outputs": [
            {
                "relative_path": "data/source_data/figure_01/F1_C.csv",
                "sha256": digest,
                "bytes": source.stat().st_size,
                "rows": 2,
            }
        ],
        "software": {},
        "parameters": {},
        "random_seeds": {},
        "limitations": ["Test fixture begins from processed data."],
    }
    provenance_path = tmp_path / "metadata" / "provenance" / "figure_01" / "F1_C.json"
    provenance_path.parent.mkdir(parents=True)
    provenance_path.write_text(json.dumps(provenance), encoding="utf-8")

    output = tmp_path / "build" / "Source_Data_partial.xlsx"
    first = build_source_workbook(tmp_path, output)
    first_bytes = output.read_bytes()
    second = build_source_workbook(tmp_path, output)
    assert output.read_bytes() == first_bytes
    assert first["sha256"] == second["sha256"]
    assert first["tables"] == 1
    assert "F1_C" not in first["missing_panels"]

    workbook = load_workbook(output, read_only=True, data_only=True)
    assert workbook["README"]["B2"].value.startswith("PARTIAL")
    assert workbook["INDEX"].max_row == 2


def _sheets_of_file(path: Path) -> set[str]:
    """Return the table names a Source Data file carries."""
    if path.suffix == ".zip":
        with zipfile.ZipFile(path) as archive:
            names = {Path(name).stem for name in archive.namelist()}
    else:
        workbook = load_workbook(path, read_only=True)
        names = set(workbook.sheetnames)
        workbook.close()
    return names - set(RESERVED_SHEETS)


def test_per_figure_files_cover_every_combined_table(source_project: Path) -> None:
    combined = source_project / "build" / "source_data" / "Combined.xlsx"
    combined_result = build_source_workbook(source_project, combined)
    per_figure = build_source_data_files(
        source_project, source_project / "build" / "source_data" / "submission"
    )

    workbook = load_workbook(combined, read_only=True)
    combined_sheets = set(workbook.sheetnames) - set(RESERVED_SHEETS)
    workbook.close()

    assert len(per_figure["files"]) == len(FIGURE_LABELS)
    assert per_figure["tables"] == combined_result["tables"]

    seen: list[str] = []
    for entry in per_figure["files"]:
        path = source_project / entry["file"]
        assert path.exists()
        sheets = _sheets_of_file(path)
        assert sheets == set(entry["sheets"])
        seen.extend(sheets)

    # Every table appears in exactly one per-figure file.
    assert sorted(seen) == sorted(combined_sheets)
    assert len(seen) == len(set(seen)) == combined_result["tables"]


def test_per_figure_files_are_self_describing(source_project: Path) -> None:
    per_figure = build_source_data_files(
        source_project, source_project / "build" / "source_data" / "submission"
    )
    entry = next(item for item in per_figure["files"] if item["figure"] == "F1")
    path = source_project / entry["file"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    readme = {row[0].value: row[1].value for row in workbook["README"].iter_rows()}
    dictionary = pd.DataFrame(
        list(workbook["DATA_DICTIONARY"].values)[1:],
        columns=list(workbook["DATA_DICTIONARY"].values)[0],
    )
    workbook.close()

    assert readme["Source Data file"] == "Source Data Figure 1"
    assert readme["Panels in this file"] == "F1_C"
    assert readme["Provenance F1_C"] == "metadata/provenance/f1/F1_C.json"
    assert "micrometre per second" in readme["Units"]
    assert set(dictionary["sheet"]) == set(entry["sheets"])
    assert dictionary.loc[dictionary["column"] == "speed_um_s", "unit"].iloc[0] == "um/s"


def test_per_figure_files_are_deterministic(source_project: Path) -> None:
    target = source_project / "build" / "source_data" / "submission"
    first = build_source_data_files(source_project, target)
    digests = {entry["file"]: entry["sha256"] for entry in first["files"]}
    second = build_source_data_files(source_project, target)
    assert {entry["file"]: entry["sha256"] for entry in second["files"]} == digests


def test_oversized_figure_falls_back_to_a_zip_archive(source_project: Path) -> None:
    target = source_project / "build" / "source_data" / "submission"
    build_source_data_files(source_project, target)
    assert (target / "Source_Data_Figure_1.xlsx").exists()

    result = build_source_data_files(source_project, target, max_bytes=1)
    entry = next(item for item in result["files"] if item["figure"] == "F1")
    archive_path = source_project / entry["file"]
    assert entry["format"] == "zip"
    assert archive_path.suffix == ".zip"
    # The superseded workbook does not linger next to the archive.
    assert not (target / "Source_Data_Figure_1.xlsx").exists()

    folder = "Source_Data_Figure_1"
    with zipfile.ZipFile(archive_path) as archive:
        names = archive.namelist()
        # Every member sits in one folder, so unzipping stays tidy.
        assert all(name.startswith(f"{folder}/") for name in names)
        assert names[:3] == [
            f"{folder}/README.txt",
            f"{folder}/INDEX.txt",
            f"{folder}/DATA_DICTIONARY.txt",
        ]
        table = pd.read_csv(
            io.BytesIO(archive.read(f"{folder}/{entry['sheets'][0]}.txt")), sep="\t"
        )
    assert list(table.columns) == ["condition", "speed_um_s"]
    assert len(table) == 2


def test_legacy_duplicate_is_dropped_only_when_it_is_redundant(
    source_project: Path,
) -> None:
    combined = source_project / "build" / "source_data" / "Combined.xlsx"
    result = build_source_workbook(source_project, combined)

    dropped = result["superseded_duplicates"]
    assert len(dropped) == 1
    assert dropped[0]["panel_id"] == "F2_A"
    assert dropped[0]["dropped_path"] == "data/source_data/figure_02/F2_X.csv"
    assert dropped[0]["retained_path"] == (
        "build/source_data/Figure_2/A/F2_A_source_data.csv"
    )
    assert dropped[0]["dropped_columns"] == 2
    assert dropped[0]["retained_columns"] == 3

    workbook = load_workbook(combined, read_only=True, data_only=True)
    sheets = set(workbook.sheetnames)
    workbook.close()
    # No sheet carries the retired panel name.
    assert not any("F2_X" in sheet for sheet in sheets)
    assert "F2_A_source_data" in sheets


def test_a_legacy_copy_with_extra_rows_survives(source_project: Path) -> None:
    """A legacy copy is kept when the current copy does not hold all its rows."""
    _write_panel(
        source_project,
        "F2_A",
        [
            (
                "data/source_data/figure_02/F2_X.csv",
                pd.DataFrame({"cell_id": [1, 2, 3], "x_um": [3.5, 4.5, 5.5]}),
            ),
            (
                "build/source_data/Figure_2/A/F2_A_source_data.csv",
                pd.DataFrame({"cell_id": [1, 2], "x_um": [3.5, 4.5]}),
            ),
        ],
    )
    combined = source_project / "build" / "source_data" / "Combined.xlsx"
    result = build_source_workbook(source_project, combined)
    assert result["superseded_duplicates"] == []
    # F1_C 2, F2_A 2 (both kept), S3_B 1, S4_A obstacles 1, plus 2 derived summaries.
    assert result["tables"] == 8


def test_sheet_names_are_readable_unique_and_untruncated(source_project: Path) -> None:
    combined = source_project / "build" / "source_data" / "Combined.xlsx"
    build_source_workbook(source_project, combined)
    workbook = load_workbook(combined, read_only=True, data_only=True)
    index = pd.DataFrame(
        list(workbook["INDEX"].values)[1:], columns=list(workbook["INDEX"].values)[0]
    )
    sheets = set(workbook.sheetnames)
    workbook.close()

    assert {"F1_C_distribution", "F1_C_statistics", "F2_A_source_data"} <= sheets
    # The panel ID appears once, not twice.
    assert not any(sheet.count("F1_C") > 1 for sheet in sheets)
    assert index["sheet"].is_unique
    assert all(len(sheet) <= 31 for sheet in index["sheet"])
    # Nothing is truncated here, so the short and full names agree.
    assert list(index["sheet"]) == list(index["table_name"])


def test_deposited_trajectories_leave_the_source_data(source_project: Path) -> None:
    """The trajectory table is deposited, not submitted, and nothing goes missing."""
    combined = source_project / "build" / "source_data" / "Combined.xlsx"
    result = build_source_workbook(source_project, combined)

    deposited = result["deposited_tables"]
    assert len(deposited) == 1
    assert deposited[0]["panel_id"] == "S4_A"
    assert deposited[0]["source_path"] in DEPOSITED_TABLE_PATHS

    workbook = load_workbook(combined, read_only=True, data_only=True)
    sheets = set(workbook.sheetnames)
    index = pd.DataFrame(
        list(workbook["INDEX"].values)[1:], columns=list(workbook["INDEX"].values)[0]
    )
    readme = {row[0]: row[1] for row in workbook["README"].iter_rows(values_only=True)}
    workbook.close()

    # The raw trajectories are gone; the summaries that replace them are present.
    assert not any("trajector" in sheet for sheet in sheets)
    assert {"S4_condition_summary", "S4_cell_summary"} <= sheets
    # INDEX says which tables are files and which are derived.
    derived = index[index["origin"] == ORIGIN_DERIVED]
    assert set(derived["table_name"]) == {"S4_condition_summary", "S4_cell_summary"}
    assert all(
        path in derived.iloc[0]["source_path"] for path in [deposited[0]["source_path"]]
    )
    assert (index["origin"] == ORIGIN_FILE).sum() == len(index) - 2
    # The README tells the reader what left and how to get it back.
    assert "Deposited, not in this file" in readme
    assert deposited[0]["sha256"] in readme["Deposited S4_A"]
    assert "build_s4.py" in readme["How to regenerate them"]


def test_deposit_summaries_carry_the_checkable_numbers(source_project: Path) -> None:
    per_figure = build_source_data_files(
        source_project, source_project / "build" / "source_data" / "submission"
    )
    entry = next(item for item in per_figure["files"] if item["figure"] == "S4")
    workbook = load_workbook(source_project / entry["file"], read_only=True, data_only=True)
    condition = pd.DataFrame(
        list(workbook["S4_condition_summary"].values)[1:],
        columns=list(workbook["S4_condition_summary"].values)[0],
    )
    cells = pd.DataFrame(
        list(workbook["S4_cell_summary"].values)[1:],
        columns=list(workbook["S4_cell_summary"].values)[0],
    )
    workbook.close()

    # The model inputs and seeds the legend quotes.
    row = condition.iloc[0]
    assert row["panel_id"] == "S4_A"
    assert row["run_speed_um_s"] == 19.9
    assert row["panel_seed"] == 17
    assert row["obstacle_seed"] == 317
    assert row["n_cells"] == 2
    # One obstacle of radius 2 um in a 148 x 96 um box.
    assert row["obstacle_count"] == 1
    assert row["obstacle_area_fraction"] == pytest.approx(
        3.141592653589793 * 4 / (148.0 * 96.0)
    )
    # Half the cells are motile in the fixture.
    assert row["realized_motile_fraction"] == pytest.approx(0.5)
    assert row["n_final_run"] == 1
    assert row["n_final_non_motile"] == 1

    # One row per simulated cell, with the numbers behind each drawn track.
    assert len(cells) == 2
    # Cell 1 is the non-motile one; it drifts 0 -> 2 -> 4 um along x over 1 s.
    drifting = cells[cells["cell_id"] == 1].iloc[0]
    assert not drifting["is_motile"]
    assert drifting["net_displacement_um"] == pytest.approx(4.0)
    assert drifting["track_path_length_um"] == pytest.approx(4.0)
    assert drifting["mean_track_speed_um_s"] == pytest.approx(4.0)
    assert drifting["final_state_label"] == "non_motile"
    # Cell 0 is the motile one, at half that speed. The motile-only aggregate
    # ignores the drifting cell, which is what run_speed_um_s compares against.
    assert cells[cells["cell_id"] == 0].iloc[0]["mean_track_speed_um_s"] == pytest.approx(2.0)
    assert row["median_track_speed_motile_um_s"] == pytest.approx(2.0)


def test_deposit_bundle_is_complete_and_verifiable(source_project: Path) -> None:
    result = build_source_data_files(
        source_project,
        source_project / "build" / "source_data" / "submission",
        deposit_dir=source_project / "build" / "source_data" / "deposit",
    )
    deposit = result["deposit"]
    bundle = source_project / deposit["bundle"]
    assert bundle.name == DEPOSIT_NAME
    names = {entry["file"] for entry in deposit["files"]}
    assert {"README.txt", "MANIFEST.tsv"} <= names
    assert deposit["tables"] == 1
    assert deposit["doi"] == "[repository and DOI pending]"

    manifest = pd.read_csv(bundle / "MANIFEST.tsv", sep="\t")
    assert list(manifest["panel_id"]) == ["S4_A"]
    assert manifest.iloc[0]["panel_seed"] == 17
    assert "plot.py" in manifest.iloc[0]["regeneration_command"]

    # The copy is byte-for-byte, so its checksum still matches the provenance.
    copied = bundle / manifest.iloc[0]["file"]
    original = source_project / next(
        path for path in DEPOSITED_TABLE_PATHS if path.endswith(manifest.iloc[0]["file"])
    )
    assert copied.read_bytes() == original.read_bytes()
    assert hashlib.sha256(copied.read_bytes()).hexdigest() == manifest.iloc[0]["sha256"]


def test_a_long_table_name_is_shortened_on_a_word_boundary() -> None:
    used: set[str] = set()
    full = _table_name("F7_D", Path("Figure_7D_effective_diffusivity_decomposition.csv"))
    assert full == "F7_D_effective_diffusivity_decomposition"
    assert _sheet_name(full, used) == "F7_D_effective_diffusivity"

    used = set()
    full = _table_name("S5_A", Path("S5_A_paired_centroid_differences.csv.gz"))
    assert full == "S5_A_paired_centroid_differences"
    assert _sheet_name(full, used) == "S5_A_paired_centroid"


def test_table_name_strips_either_panel_naming_convention() -> None:
    assert _table_name("F1_C", Path("F1_C_distribution.csv")) == "F1_C_distribution"
    assert _table_name("F6_A", Path("Figure_6A_summary_95ci.csv")) == "F6_A_summary_95ci"
    assert _table_name("F4_A", Path("schematic_asset.csv")) == "F4_A_schematic_asset"
    assert _table_name("S1_A", Path("S1_A.csv.gz")) == "S1_A_table"


def test_column_unit_reads_the_column_name() -> None:
    assert column_unit("speed_um_s") == "um/s"
    assert column_unit("speed_um_s_ci95_high") == "um/s"
    assert column_unit("x_um") == "um"
    assert column_unit("AnTc_ng_per_mL") == "ng/mL"
    assert column_unit("time_s") == "s"
    assert column_unit("growth_rate_per_h") == "1/h"
    assert column_unit("growth_penalty_percent") == "%"
    assert column_unit("mass_fraction") == "fraction (0-1)"
    assert column_unit("_window_start_min") == "min"
    # "min" here means minimum, not minutes; the builder states no unit.
    assert column_unit("grid_speed_min") == ""
    assert column_unit("condition") == ""
